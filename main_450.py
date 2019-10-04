import time
import os
import glob
import numpy as np
import random
import re
from NN_Model_480 import createNeuralNetworkModel
from Env import runSimulation, runSimulation_input, activitySequence, activity
import multiprocessing as mp
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side


t_start = time.time()

# user defined parameters
# problem parameters
timeDistribution = "deterministic"    # deterministic, exponential, uniform_1, uniform_2, ...

# CPU parameters
numberOfCpuProcessesToGenerateData = 16   # paoloPC has 16 cores
maxTasksPerChildToGenerateData = 4        # 4 is the best for paoloPC


# input state vector  parameters
numberOfActivitiesInStateVector = 6
rescaleFactorTime = 0.1
#timeHorizon = 10

# random generation parameters
numberOfSimulationRunsToGenerateData =2000
numberOfSimulationRunsToTestPolicy = 1

# train parameters
percentageOfFilesTest = 0.1
importExistingNeuralNetworkModel = False
neuralNetworkModelAlreadyExists = False
numberOfEpochs = 20 #walk entire samples
learningRate = 0.0005

# paths
relativePath = os.path.dirname(__file__)
absolutePathProjects = relativePath + "/RG_450/"


# other parameters
np.set_printoptions(precision=4)    # print precision of numpy variables

# initialise variables
numberOfActivities = None
numberOfResources = None
activitySequences = []
decisions_indexActivity = []
decisions_indexActivityPowerset = []
states = []
actions = []
sumTotalDurationRandomTestRecord = []
sumTotalDurationWithNeuralNetworkModelTestRecord = []
sumTotalDurationWithCriticalResourceTestRecord = []
sumTotalDurationWithShortestProcessingTestRecord = []
sumTotalDurationWithShortestSumDurationTestRecord = []
sumTotalDurationRandomTrainRecord = []
sumTotalDurationWithNeuralNetworkModelTrainRecord = []
sumTotalDurationWithCriticalResourceTrainRecord = []
sumTotalDurationWithShortestProcessingTrainRecord = []
sumTotalDurationWithShortestSumDurationTrainRecord = []

# read all activity sequences from database
absolutePathProjectsGlob = absolutePathProjects + "*.txt"
files = sorted(glob.glob(absolutePathProjectsGlob))

# divide all activity sequences in training and test set
numberOfFiles = len(files) #450
numberOfFilesTest = round(numberOfFiles * percentageOfFilesTest) #45
numberOfFilesTrain = numberOfFiles - numberOfFilesTest #405
indexFiles = list(range(0, numberOfFiles))
indexFilesTrain = []
indexFilesTest = []


# choose the last element of every set to test
for i in range(1,numberOfFilesTest+1):
    indexFilesTest.append(indexFiles[i*10-1])

for i in indexFiles:
    if i not in indexFilesTest:
        indexFilesTrain.append(i)

# organize the read activity sequences in classes
for i in range(numberOfFiles):
    file = files[i]
    #print(file)
    # create a new activitySequence object
    currentActivitySequence = activitySequence()
    with open(file,"r") as f:
        currentActivitySequence.index = i
        currentActivitySequence.fileName = os.path.basename(f.name)
        #print(currentActivitySequence.fileName)
        # allLines = f.read()
        # print(allLines)
        next(f)
        firstLine = f.readline()    # information about numberOfActivities and numberOfResourceTypes
        firstLineDecomposed = re.split(" +", firstLine)
        numberOfActivities = (int(firstLineDecomposed[1])-2)    # the first and last dummy activity do not count
        currentActivitySequence.numberOfActivities = numberOfActivities
        # print("numberOfActivities = " + str(currentActivitySequence.numberOfActivities))
        secondLine = f.readline()   # information about total available resources
        secondLineDecomposed = re.split(" +", secondLine)
        #print(secondLineDecomposed)
        del secondLineDecomposed[0]
        secondLineDecomposed[-1]=secondLineDecomposed[-1].strip('\n')#only string have attributes of strip and delete '\n' at the end
        #print(secondLineDecomposed)
        numberOfResources = 0
        # print(len(secondLineDecomposed))
        #secondLineDecomposed=[int(secondLineDecomposed)]
        secondLineDecomposed = list(map(int,secondLineDecomposed))
        #print(secondLineDecomposed)
        for totalResources in secondLineDecomposed:
            numberOfResources += 1
            #print(numberOfResources)
            currentActivitySequence.totalResources.append(int(totalResources))
            #print(currentActivitySequence.totalResources)
        currentActivitySequence.numberOfResources = numberOfResources
        next(f)
        thirdLine = f.readline()   # information about starting activities
        thirdLineDecomposed = re.split(" +", thirdLine)
        for IdActivity in thirdLineDecomposed[7:-1]:
            currentActivitySequence.indexStartActivities.append(int(IdActivity)-2)
        #print("indexStartingActivities = " + str(currentActivitySequence.indexStartActivities))
        line = f.readline()
        while line:
            #print(line, end="")
            lineDecomposed = re.split(" +", line)
            if int(lineDecomposed[1]) == 0:
                break
            else:
                currentActivity = activity()
                currentActivity.time = int(lineDecomposed[1])
                currentActivity.requiredResources = [ int(lineDecomposed[2]),int(lineDecomposed[3]),int(lineDecomposed[4]),int(lineDecomposed[5])]
                for IdFollowingActivity in lineDecomposed[7:-1]:
                    if int(IdFollowingActivity) != numberOfActivities+2:    #if the following action is not the last dummy activity
                        currentActivity.indexFollowingActivities.append(int(IdFollowingActivity) - 2)
            currentActivitySequence.activities.append(currentActivity)
            line = f.readline()
        #add indexes to list of activities
        for j in range(len(currentActivitySequence.activities)):
            currentActivitySequence.activities[j].index = j
        #add numberOfPreviousActivities
        for Activity in currentActivitySequence.activities:
            for IndexFollowingActivity in Activity.indexFollowingActivities:
                currentActivitySequence.activities[IndexFollowingActivity].numberOfPreviousActivities += 1
    activitySequences.append(currentActivitySequence)

#stateVectorLength = numberOfActivitiesInStateVector + numberOfActivitiesInStateVector * numberOfResources + numberOfResources+ timeHorizon * numberOfResources
stateVectorLength = numberOfActivitiesInStateVector + numberOfActivitiesInStateVector * numberOfResources + numberOfResources

# compute decisions: each decision corresponds to a start of an activity in the local reference system (more than one decision can be taken at once)
for i in range(0,numberOfActivitiesInStateVector):
    decisions_indexActivity.append(i)


#--------------------------------------------------------------RANDOM-----------------------------------------------------------------------------
####  GENERATE TRAINING DATA USING RANDOM DECISIONS (WITHOUT USING pool.map) ####
print('######  RANDOM DECISION ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "generateData"
    currentRunSimulation_input.randomDecisionProbability = 1
    currentRunSimulation_input.policyType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences and append training data

for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
    activitySequences[indexFilesTrain[i]].totalDurationStandardDeviation = runSimulation_outputs[i].totalDurationStDev
    activitySequences[indexFilesTrain[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
    activitySequences[indexFilesTrain[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
    activitySequences[indexFilesTrain[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
    activitySequences[indexFilesTrain[i]].trivialDecisionPercentageMean = runSimulation_outputs[i].trivialDecisionPercentageMean

    for currentStateActionPair in runSimulation_outputs[i].stateActionPairsOfBestRun:
        states.append(currentStateActionPair.state)
        actions.append(currentStateActionPair.action)


#correspondence best states and actions pairs --> len(states) = len(actions)
#print('state',states)
#print('actions:',actions)


####  TRAIN MODEL USING TRAINING DATA  ####
# look for existing model
if importExistingNeuralNetworkModel:
    print("check if a neural network model exists")
    if neuralNetworkModelAlreadyExists:
        print("import neural network model exists")

    else:
        neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
        # neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actions[0]))
else:
    neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
    # neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actions[0]))

neuralNetworkModel.fit({"input": states}, {"targets": actions}, n_epoch=numberOfEpochs, snapshot_step=500,show_metric=True,batch_size=32,validation_set=0.2)
# output = neuralNetworkModel.predict(states)
# print(output)

####  CREATE BENCHMARK WITH RANDOM DECISIONS ALSO WITH TEST ACTIVITY SEQUENCES  ####
print('######  RANDOM DECISION ON TEST ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 1
    currentRunSimulation_input.policyType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences

for i in range(numberOfFilesTest):
    activitySequences[indexFilesTest[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
    activitySequences[indexFilesTest[i]].totalDurationStandardDeviation = runSimulation_outputs[i].totalDurationStDev
    activitySequences[indexFilesTest[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
    activitySequences[indexFilesTest[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
    activitySequences[indexFilesTest[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
    activitySequences[indexFilesTest[i]].trivialDecisionPercentageMean = runSimulation_outputs[i].trivialDecisionPercentageMean


#-----------------------------------------------------------------NN------------------------------------------------------------------------------
####  TEST NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ####
# run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
print('###### NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "neuralNetworkModel"
    currentRunSimulation_input.decisionTool = neuralNetworkModel
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTrain[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean



####  TEST NEURAL NETWORK MODEL ON TEST ACTIVITY SEQUENCES  ####
# run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
print('###### NEURAL NETWORK MODEL ON TEST ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "neuralNetworkModel"
    currentRunSimulation_input.decisionTool = neuralNetworkModel
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTest[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

# ---------------------------------------------------------Critical Resource----------------------------------------------------------------------------
####  TEST CRITICAL RESOURCE METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### CRITICAL RESOURCE METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "most critical resource"
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithCriticalResource = runSimulation_outputs[i].totalDurationMean

####  TEST CRITICAL RESOURCE METHOD ON TEST ACTIVITY SEQUENCES  ####
print('###### CRITICAL RESOURCE METHOD ON TEST ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "most critical resource"
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTest[i]].totalDurationWithCriticalResource = currentRunSimulation_output.totalDurationMean

# ---------------------------------------------------------Shortest Processing Time----------------------------------------------------------------------------
####  TEST SHORTEST PROCESSING TIME METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### SHORTEST PROCESSING TIME METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest processing time"
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithShortestProcessingTime = runSimulation_outputs[i].totalDurationMean

####  TEST SHORTEST PROCESSING TIME METHOD ON TEST ACTIVITY SEQUENCES  ####
print('###### SHORTEST PROCESSING TIME METHOD ON TEST ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest processing time"
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTest[i]].totalDurationWithShortestProcessingTime = currentRunSimulation_output.totalDurationMean


# ---------------------------------------------------------shortest sumDuration including successor----------------------------------------------------------------------------
####  TEST SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest sumDuration including successor"
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithShortestSumDuration = runSimulation_outputs[i].totalDurationMean

####  TEST SHORTEST SUMDURATION INCLUDING SUCCESSOR TIME METHOD ON TEST ACTIVITY SEQUENCES  ####
print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TEST ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest sumDuration including successor"
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    #currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTest[i]].totalDurationWithShortestSumDuration = currentRunSimulation_output.totalDurationMean


# ------------------------------------------------------EVALUATION-----------------------------------------------------------------------------
####  EVALUATION OF RESULTS OF TRAIN ACTIVITY SEQUENCES  ####
sumTotalDurationRandomTrain = 0
sumTotalDurationWithCriticalResourceTrain = 0
sumTotalDurationWithShortestProcessingTrain = 0
sumTotalDurationWithNeuralNetworkModelTrain = 0
sumTotalDurationWithShortestSumDurationTrain = 0

for i in range(numberOfFilesTrain):
    sumTotalDurationRandomTrain += activitySequences[indexFilesTrain[i]].totalDurationMean
    sumTotalDurationRandomTrain = round(sumTotalDurationRandomTrain,4)
    sumTotalDurationWithNeuralNetworkModelTrain += activitySequences[indexFilesTrain[i]].totalDurationWithPolicy
    sumTotalDurationWithCriticalResourceTrain += activitySequences[indexFilesTrain[i]].totalDurationWithCriticalResource
    sumTotalDurationWithShortestProcessingTrain += activitySequences[indexFilesTrain[i]].totalDurationWithShortestProcessingTime
    sumTotalDurationWithShortestSumDurationTrain += activitySequences[indexFilesTrain[i]].totalDurationWithShortestSumDuration

sumTotalDurationRandomTrainRecord.append(sumTotalDurationRandomTrain)
sumTotalDurationWithNeuralNetworkModelTrainRecord.append(sumTotalDurationWithNeuralNetworkModelTrain)
sumTotalDurationWithCriticalResourceTrainRecord.append(sumTotalDurationWithCriticalResourceTrain)
sumTotalDurationWithShortestProcessingTrainRecord.append(sumTotalDurationWithShortestProcessingTrain)
sumTotalDurationWithShortestSumDurationTrainRecord.append(sumTotalDurationWithShortestSumDurationTrain)

####  EVALUATION OF NN RESULTS OF TEST ACTIVITY SEQUENCES  ####
sumTotalDurationRandomTest = 0
sumTotalDurationWithNeuralNetworkModelTest = 0
sumTotalDurationWithCriticalResourceTest = 0
sumTotalDurationWithShortestProcessingTest = 0
sumTotalDurationWithShortestSumDurationTest = 0

for i in range(numberOfFilesTest):
    sumTotalDurationRandomTest += activitySequences[indexFilesTest[i]].totalDurationMean
    sumTotalDurationRandomTest = round(sumTotalDurationRandomTest,4)
    sumTotalDurationWithNeuralNetworkModelTest += activitySequences[indexFilesTest[i]].totalDurationWithPolicy
    sumTotalDurationWithCriticalResourceTest += activitySequences[indexFilesTest[i]].totalDurationWithCriticalResource
    sumTotalDurationWithShortestProcessingTest += activitySequences[indexFilesTest[i]].totalDurationWithShortestProcessingTime
    sumTotalDurationWithShortestSumDurationTest += activitySequences[indexFilesTest[i]].totalDurationWithShortestSumDuration


sumTotalDurationRandomTestRecord.append(sumTotalDurationRandomTest)
sumTotalDurationWithNeuralNetworkModelTestRecord.append(sumTotalDurationWithNeuralNetworkModelTest)
sumTotalDurationWithCriticalResourceTestRecord.append(sumTotalDurationWithCriticalResourceTest)
sumTotalDurationWithShortestProcessingTestRecord.append(sumTotalDurationWithShortestProcessingTest)
sumTotalDurationWithShortestSumDurationTestRecord.append(sumTotalDurationWithShortestSumDurationTest)



print("sumTotalDurationRandomTrain = " + str(sumTotalDurationRandomTrain))
print("sumTotalDurationWithNeuralNetworkModelTrain = " + str(sumTotalDurationWithNeuralNetworkModelTrain))
print("sumTotalDurationWithCriticalResourceTrain = " + str(sumTotalDurationWithCriticalResourceTrain))
print("sumTotalDurationWithShortestProcessingTrain = " + str(sumTotalDurationWithShortestProcessingTrain))
print("sumTotalDurationWithShortestSumDurationTrain = " + str(sumTotalDurationWithShortestSumDurationTrain))
print("sumTotalDurationRandomTest = " + str(sumTotalDurationRandomTest))
print("sumTotalDurationWithNeuralNetworkModelTest = " + str(sumTotalDurationWithNeuralNetworkModelTest))
print("sumTotalDurationWithCriticalResourceTest = " + str(sumTotalDurationWithCriticalResourceTest))
print("sumTotalDurationWithShortestProcessingTest = " + str(sumTotalDurationWithShortestProcessingTest))
print("sumTotalDurationWithShortestSumDurationTest = " + str(sumTotalDurationWithShortestSumDurationTest))


# compute computation time
t_end = time.time()
t_computation = t_end - t_start
print("t_computation = " + str(t_computation))

#
# #write ouput to excel
# wb = Workbook()
# ws = wb.create_sheet('RG30_duration',0)
#
# #combine rows
# ws.merge_cells('A1:B1')
# ws.merge_cells('D1:G1')
# ws.merge_cells('K1:N1')
#
# #name it
# ws['A1'] = 'number of simulation runs'
# # ws['A2'] = 'Prob[number of ready to start activity]'
# ws['B2'] = 'train Topology name'
# ws['J2'] = 'test Topology name'
# ws['C1'] = numberOfSimulationRunsToGenerateData
# ws['D1'] = 'train Solution random'
# ws['H1'] = 'train policy'
# ws['I1'] = 'train heuristic'
# ws['K1'] = 'test Solution random'
# ws['O1'] = 'test policy'
# ws['P1'] = 'test heuristic'
# ws['A3'] = 'computation time'
# ws['Q1'] = 'sumRandomTr'
# ws['R1'] = 'sumNNTr'
# ws['S1'] = 'sumHeuristicTr'
# ws['T1'] = 'sumRandomTe'
# ws['U1'] = 'sumNNTe'
# ws['V1'] = 'sumHeuristicTe'
#
# #Train data
# ws['D2'] = 'E[T]'
# ws['E2'] = 'StDev[T]'
# ws['F2'] = 'Min[T]'
# ws['G2'] = 'Max[T]'
# ws['H2'] = '[T]'
# ws['I2'] = '[T]'
#
# #Test data
# ws['K2'] = 'E[T]'
# ws['L2'] = 'StDev[T]'
# ws['M2'] = 'Min[T]'
# ws['N2'] = 'Max[T]'
# ws['O2'] = '[T]'
# ws['P2'] = '[T]'
#
#
#
# #change column width and height
# ws.column_dimensions['A'].width = 17.0
# ws.column_dimensions['B'].width = 11.0
# ws.column_dimensions['J'].width = 11.0
# ws.column_dimensions['H'].width = 11.0
# ws.column_dimensions['O'].width = 11.0
# ws.column_dimensions['P'].width = 13.0
# ws.column_dimensions['I'].width = 13.0
# ws.column_dimensions['Q'].width = 14.0
# ws.column_dimensions['R'].width = 14.0
# ws.column_dimensions['S'].width = 14.0
# ws.column_dimensions['T'].width = 14.0
# ws.column_dimensions['U'].width = 14.0
# ws.column_dimensions['V'].width = 14.0
# ws.row_dimensions[2].height = 45
# ws.row_dimensions[1].height = 30
#
# #alignment can be accessed only per cell
# align = Alignment(horizontal='center',vertical='center',wrap_text=True)
# ws['A1'].alignment = align
# ws['D1'].alignment = align
# ws['K1'].alignment = align
# ws['H1'].alignment = align
# ws['O1'].alignment = align
# ws['P1'].alignment = align
# ws['I1'].alignment = align
# for item in ws['A2:V2'][0]:
#     item.alignment = align
#
# for item in ws['Q1:V1'][0]:
#     item.alignment = align
#
# # ws.cell(row=len_probabilityDistributionNumberOfReadyToStartActivities+3, column=1).value = "computation time"
# # ws.cell(row=len_probabilityDistributionNumberOfReadyToStartActivities+4, column=1).value = t_computation
# for i in range(numberOfFilesTrain):
#     ws.cell(row=i+3, column=2).value = activitySequences[indexFilesTrain[i]].fileName[:-4]
#     ws.cell(row=i+3, column=4).value = activitySequences[indexFilesTrain[i]].totalDurationMean
#     ws.cell(row=i+3, column=5).value = activitySequences[indexFilesTrain[i]].totalDurationStandardDeviation
#     ws.cell(row=i+3, column=6).value = activitySequences[indexFilesTrain[i]].totalDurationMin
#     ws.cell(row=i+3, column=7).value = activitySequences[indexFilesTrain[i]].totalDurationMax
#     #using NN_Model results
#     ws.cell(row=i + 3, column=8).value = activitySequences[indexFilesTrain[i]].totalDurationWithPolicy
#     ws.cell(row=i + 3, column=9).value = activitySequences[indexFilesTrain[i]].totalDurationWithHeuristic
#
# for i in range(numberOfFilesTest):
#     ws.cell(row=i + 3, column=10).value = activitySequences[indexFilesTest[i]].fileName[:-4]
#     ws.cell(row=i + 3, column=11).value = activitySequences[indexFilesTest[i]].totalDurationMean
#     ws.cell(row=i + 3, column=12).value = activitySequences[indexFilesTest[i]].totalDurationStandardDeviation
#     ws.cell(row=i + 3, column=13).value = activitySequences[indexFilesTest[i]].totalDurationMin
#     ws.cell(row=i + 3, column=14).value = activitySequences[indexFilesTest[i]].totalDurationMax
#     # using NN_Model results
#     ws.cell(row=i + 3, column=15).value = activitySequences[indexFilesTest[i]].totalDurationWithPolicy
#     ws.cell(row=i + 3, column=16).value = activitySequences[indexFilesTest[i]].totalDurationWithHeuristic
#
# ws.cell(row=2, column=17).value = sumTotalDurationRandomTrain
# ws.cell(row=2, column=18).value = sumTotalDurationWithNeuralNetworkModelTrain
# ws.cell(row=2, column=19).value = sumTotalDurationWithHeuristicTrain
# ws.cell(row=2, column=20).value = sumTotalDurationRandomTest
# ws.cell(row=2, column=21).value = sumTotalDurationWithNeuralNetworkModelTest
# ws.cell(row=2, column=22).value = sumTotalDurationWithHeuristicTest
#
#
# ws.cell(row=4, column=1).value = round(t_computation,2)
#
# wb.save(relativePath + "/database_RG30_Newdata/times2000_epochs50.xlsx")
#
